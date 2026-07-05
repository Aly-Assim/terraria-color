from pathlib import Path
from urllib.parse import urljoin, urlparse, unquote, quote
import csv
import hashlib
import re
import shutil
import sqlite3
import time

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


BASE_URL = "https://terraria.wiki.gg"

EXCEL_PATH = Path("source/all_items_terraria_145.xlsx")

DATA_DIR = Path("data")
DEBUG_DIR = Path("debug")

CACHE_PAGES_DIR = Path("cache/pages")
CACHE_IMAGES_DIR = Path("cache/images")

IMAGES_DIR = Path("images")
INVENTORY_IMAGES_DIR = IMAGES_DIR / "inventory"
WORLD_IMAGES_DIR = IMAGES_DIR / "world"
COLOR_IMAGES_DIR = IMAGES_DIR / "color"

DB_PATH = DATA_DIR / "terraria_blocks.db"

REPORT_PATH = DATA_DIR / "report.txt"
EXCEL_READ_CSV_PATH = DATA_DIR / "excel_blocks_read.csv"
ALL_ENTRIES_CSV_PATH = DATA_DIR / "all_entries.csv"
MANUAL_FIXES_CSV_PATH = DATA_DIR / "manual_fixes.csv"
PROBLEM_ENTRIES_CSV_PATH = DATA_DIR / "problem_entries.csv"

REQUEST_DELAY = 0.20

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0 Safari/537.36"
    )
}

BAD_NAME_PARTS = [
    "desktop",
    "console",
    "mobile",
    "old-gen",
    "old gen",
    "version",
    "versions",
    "file:",
    "special:",
    "category:",
    "template:",
    "help:",
    "view or edit",
    "edit",
    "size=",
    "tiles ",
]

BAD_IMAGE_PARTS = [
    "desktop",
    "console",
    "mobile",
    "old-gen",
    "old_gen",
    "3ds",
    "journey",
    "classic",
    "expert",
    "master",
    "rarity",
    "research",
    "check",
    "cross",
    "yes",
    "no",
    "icon",
    "logo",
]


def setup_dirs():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    DEBUG_DIR.mkdir(parents=True, exist_ok=True)

    CACHE_PAGES_DIR.mkdir(parents=True, exist_ok=True)
    CACHE_IMAGES_DIR.mkdir(parents=True, exist_ok=True)

    if IMAGES_DIR.exists():
        shutil.rmtree(IMAGES_DIR)

    INVENTORY_IMAGES_DIR.mkdir(parents=True, exist_ok=True)
    WORLD_IMAGES_DIR.mkdir(parents=True, exist_ok=True)
    COLOR_IMAGES_DIR.mkdir(parents=True, exist_ok=True)


def clean_name(text):
    if text is None:
        return None

    text = str(text).strip()

    if not text:
        return None

    text = text.replace(".png", "")
    text = text.replace(".gif", "")
    text = text.replace(".jpg", "")
    text = text.replace(".jpeg", "")
    text = text.replace(".webp", "")

    text = text.replace("(placed)", "")
    text = text.replace("(Placed)", "")
    text = text.replace("_", " ")

    text = re.sub(r"\s+", " ", text)

    return text.strip()


def split_camel_case(text):
    if not text:
        return text

    text = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", text)
    text = re.sub(r"(?<=[A-Z])(?=[A-Z][a-z])", " ", text)

    return text


def normalize_name(text):
    text = clean_name(text)

    if not text:
        return None

    text = split_camel_case(text)
    text = text.lower()
    text = text.replace("’", "'")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text)

    return text.strip()


def safe_filename(text):
    if not text:
        text = "unknown"

    text = str(text).lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = text.strip("_")

    return text or "unknown"


def is_bad_name(text):
    if not text:
        return True

    text = clean_name(text)

    if not text:
        return True

    lowered = text.lower()

    for bad in BAD_NAME_PARTS:
        if bad in lowered:
            return True

    if lowered in {"true", "false", "none", "nil"}:
        return True

    if re.fullmatch(r"\d+", lowered):
        return True

    return False


def canonicalize_name(name):
    return normalize_name(name)


def get_filename_from_url(url):
    if not url:
        return None

    parsed = urlparse(url)
    path = unquote(parsed.path)
    parts = [part for part in path.split("/") if part]

    if not parts:
        return None

    filename = parts[-1]

    if "thumb" in parts and len(parts) >= 2:
        previous = parts[-2]

        if "." in previous:
            filename = previous

    return filename


def get_extension_from_url(url):
    filename = get_filename_from_url(url)

    if not filename or "." not in filename:
        return ".png"

    extension = Path(filename).suffix.lower()

    if extension in {".png", ".gif", ".jpg", ".jpeg", ".webp"}:
        return extension

    return ".png"


def page_url_from_name(name):
    page_title = clean_name(name)

    if not page_title:
        return None

    page_title = page_title.replace(" ", "_")
    return f"{BASE_URL}/wiki/{quote(page_title, safe='/')}"


def render_url_from_page_url(page_url):
    parsed = urlparse(page_url)
    path = parsed.path

    if "/wiki/" not in path:
        return page_url

    page_title = unquote(path.split("/wiki/", 1)[1])
    return f"{BASE_URL}/wiki/{quote(page_title, safe='/')}?action=render"


def get_img_src(img):
    if img is None:
        return None

    return img.get("src") or img.get("data-src")


def full_url(src):
    if not src:
        return None

    return urljoin(BASE_URL, src)


def image_text_from_img(img, url=None):
    alt = ""
    src = ""

    if img is not None:
        alt = img.get("alt", "") or ""
        src = get_img_src(img) or ""

    filename = get_filename_from_url(url or src) or ""

    return f"{alt} {src} {filename}".lower()


def image_is_bad(img, url=None):
    text = image_text_from_img(img, url)

    for bad in BAD_IMAGE_PARTS:
        if bad in text:
            return True

    return False


def image_is_placed(img, url=None):
    text = image_text_from_img(img, url)
    return "placed" in text


def image_matches_block_name(img, block_name, url=None):
    text = image_text_from_img(img, url)

    normalized_text = normalize_name(text)
    normalized_block = normalize_name(block_name)

    if not normalized_text or not normalized_block:
        return False

    compact_text = normalized_text.replace(" ", "")
    compact_block = normalized_block.replace(" ", "")

    return compact_block in compact_text


def extract_image_formula_url(value):
    if not isinstance(value, str):
        return None

    match = re.search(r'IMAGE\("([^"]+)"', value, flags=re.IGNORECASE)

    if match:
        return match.group(1)

    if value.startswith("http://") or value.startswith("https://"):
        return value

    return None


def find_category_for_column(sheet, col):
    for c in range(col, 0, -1):
        value = sheet.cell(row=1, column=c).value

        if value:
            category = clean_name(value)

            if category and not is_bad_name(category):
                return category

    return None


def read_blocks_from_excel():
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(
            f"Fichier introuvable : {EXCEL_PATH}\n"
            "Mets ton fichier Excel ici : source/all_items_terraria_145.xlsx"
        )

    workbook = load_workbook(EXCEL_PATH, data_only=False)

    if "Blocks" in workbook.sheetnames:
        sheet = workbook["Blocks"]
    else:
        candidates = [name for name in workbook.sheetnames if "block" in name.lower()]

        if not candidates:
            raise ValueError(f"Aucune feuille Blocks trouvée. Feuilles : {workbook.sheetnames}")

        sheet = workbook[candidates[0]]

    entries = []
    seen = {}

    for row in range(2, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(row=row, column=col).value

            if not isinstance(value, str):
                continue

            name = clean_name(value)

            if is_bad_name(name):
                continue

            image_cell_value = sheet.cell(row=row, column=col + 1).value
            excel_image_url = extract_image_formula_url(image_cell_value)

            # On veut vraiment les cellules "nom + image juste à droite".
            if not excel_image_url:
                continue

            canonical_name = canonicalize_name(name)

            if not canonical_name:
                continue

            category = find_category_for_column(sheet, col)

            if canonical_name in seen:
                seen[canonical_name]["duplicate_names"].append(name)
                seen[canonical_name]["duplicate_rows"].append(row)
                continue

            entry = {
                "source_index": len(entries),
                "row": row,
                "col": col,
                "name": name,
                "canonical_name": canonical_name,
                "normalized_name": normalize_name(name),
                "category_name": category,
                "excel_image_url": excel_image_url,
                "duplicate_names": [],
                "duplicate_rows": [],
            }

            entries.append(entry)
            seen[canonical_name] = entry

    return entries


def write_excel_read_csv(entries):
    with EXCEL_READ_CSV_PATH.open("w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)

        writer.writerow([
            "source_index",
            "row",
            "col",
            "name",
            "canonical_name",
            "category_name",
            "excel_image_url",
            "duplicate_names",
            "duplicate_rows",
        ])

        for item in entries:
            writer.writerow([
                item["source_index"],
                item["row"],
                item["col"],
                item["name"],
                item["canonical_name"],
                item["category_name"],
                item["excel_image_url"],
                " | ".join(item["duplicate_names"]),
                " | ".join(str(x) for x in item["duplicate_rows"]),
            ])


def download_html(url, cache_name):
    cache_path = CACHE_PAGES_DIR / cache_name

    if cache_path.exists():
        return cache_path.read_text(encoding="utf-8"), True

    response = requests.get(url, headers=HEADERS, timeout=30)
    response.raise_for_status()

    html = response.text
    cache_path.write_text(html, encoding="utf-8")

    time.sleep(REQUEST_DELAY)

    return html, False


def url_hash(url):
    return hashlib.sha1(url.encode("utf-8")).hexdigest()


def download_image_to_cache(url):
    if not url:
        return None

    extension = get_extension_from_url(url)
    cache_path = CACHE_IMAGES_DIR / f"{url_hash(url)}{extension}"

    if cache_path.exists():
        return cache_path

    response = requests.get(url, headers=HEADERS, timeout=30)
    response.raise_for_status()

    cache_path.write_bytes(response.content)

    time.sleep(REQUEST_DELAY)

    return cache_path


def copy_image_from_cache(url, output_folder, base_name):
    if not url:
        return None

    extension = get_extension_from_url(url)
    output_path = output_folder / f"{base_name}{extension}"

    cache_path = download_image_to_cache(url)
    shutil.copyfile(cache_path, output_path)

    return str(output_path)


def try_copy_image(url, output_folder, base_name):
    try:
        return copy_image_from_cache(url, output_folder, base_name), None
    except Exception as error:
        return None, str(error)


def extract_infobox(soup):
    selectors = [
        "aside.portable-infobox",
        ".portable-infobox",
        "table.infobox",
        ".infobox",
    ]

    for selector in selectors:
        infobox = soup.select_one(selector)

        if infobox is not None:
            return infobox

    return None


def clean_infobox_for_strict_parse(infobox):
    for tag_name in ["script", "style", "noscript"]:
        for tag in infobox.find_all(tag_name):
            tag.decompose()

    selectors = [
        ".mw-editsection",
        ".pi-header",
        ".pi-navigation",
        ".infobox-notice",
        ".game-icon",
        ".eicons",
        ".eil",
        ".i.s",
        ".plainlinks",
    ]

    for selector in selectors:
        for tag in infobox.select(selector):
            tag.decompose()

    return infobox


def extract_internal_ids_from_infobox(infobox):
    text = infobox.get_text(" ", strip=True)

    internal_item_id = None
    internal_tile_id = None

    item_match = re.search(r"Internal\s+Item\s+ID\s*:\s*(\d+)", text, flags=re.IGNORECASE)

    if item_match:
        internal_item_id = int(item_match.group(1))

    tile_match = re.search(r"Internal\s+Tile\s+ID\s*:\s*(\d+)", text, flags=re.IGNORECASE)

    if tile_match:
        internal_tile_id = int(tile_match.group(1))

    return internal_item_id, internal_tile_id


def extract_images_from_infobox(infobox, block_name):
    inventory_candidates = []
    world_candidates = []

    imgs = infobox.find_all("img")

    for img in imgs:
        src = get_img_src(img)

        if not src:
            continue

        url = full_url(src)

        if image_is_bad(img, url):
            continue

        if image_is_placed(img, url):
            score = 100

            if image_matches_block_name(img, block_name, url):
                score += 50

            world_candidates.append((score, url))
            continue

        if image_matches_block_name(img, block_name, url):
            score = 100

            try:
                width = int(img.get("width", "0"))
                height = int(img.get("height", "0"))
            except Exception:
                width = 0
                height = 0

            if width and height and width <= 96 and height <= 96:
                score += 20

            inventory_candidates.append((score, url))

    inventory_url = None
    world_url = None

    if inventory_candidates:
        inventory_candidates.sort(key=lambda item: item[0], reverse=True)
        inventory_url = inventory_candidates[0][1]

    if world_candidates:
        world_candidates.sort(key=lambda item: item[0], reverse=True)
        world_url = world_candidates[0][1]

    return inventory_url, world_url, len(imgs)


def parse_page_strict(block):
    name = block["name"]
    page_url = page_url_from_name(name)
    render_url = render_url_from_page_url(page_url)

    cache_name = f"page_{safe_filename(block['canonical_name'])}.html"

    html, from_cache = download_html(render_url, cache_name)

    soup = BeautifulSoup(html, "html.parser")
    infobox = extract_infobox(soup)

    result = {
        "page_url": page_url,
        "from_cache": from_cache,
        "inventory_image_url": None,
        "world_image_url": None,
        "color_image_url": None,
        "internal_item_id": None,
        "internal_tile_id": None,
        "infobox_image_count": None,
        "problem": None,
    }

    if infobox is None:
        result["problem"] = "infobox_not_found"
        return result

    infobox = clean_infobox_for_strict_parse(infobox)

    inventory_url, world_url, infobox_image_count = extract_images_from_infobox(infobox, name)
    internal_item_id, internal_tile_id = extract_internal_ids_from_infobox(infobox)

    result["inventory_image_url"] = inventory_url
    result["world_image_url"] = world_url
    result["color_image_url"] = world_url
    result["internal_item_id"] = internal_item_id
    result["internal_tile_id"] = internal_tile_id
    result["infobox_image_count"] = infobox_image_count

    missing = []

    if not inventory_url:
        missing.append("inventory_image")

    if not world_url:
        missing.append("world_image")

    if missing:
        result["problem"] = "missing_" + "_and_".join(missing)

    return result


def create_database():
    if DB_PATH.exists():
        DB_PATH.unlink()

    connection = sqlite3.connect(DB_PATH)
    cursor = connection.cursor()

    cursor.execute("""
        CREATE TABLE objects (
            local_id INTEGER PRIMARY KEY,

            object_type TEXT NOT NULL,
            is_wall INTEGER NOT NULL,

            name TEXT NOT NULL,
            normalized_name TEXT NOT NULL,
            canonical_name TEXT NOT NULL,

            category_name TEXT,
            page_url TEXT,

            inventory_image_url TEXT,
            world_image_url TEXT,
            color_image_url TEXT,

            inventory_image_path TEXT,
            world_image_path TEXT,
            color_image_path TEXT,

            internal_item_id INTEGER,
            internal_tile_id INTEGER,

            source TEXT NOT NULL,
            status TEXT NOT NULL,
            problem TEXT
        )
    """)

    connection.commit()
    return connection


def insert_object(connection, item):
    cursor = connection.cursor()

    cursor.execute("""
        INSERT INTO objects (
            local_id,

            object_type,
            is_wall,

            name,
            normalized_name,
            canonical_name,

            category_name,
            page_url,

            inventory_image_url,
            world_image_url,
            color_image_url,

            inventory_image_path,
            world_image_path,
            color_image_path,

            internal_item_id,
            internal_tile_id,

            source,
            status,
            problem
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        item["local_id"],

        "block",
        0,

        item["name"],
        item["normalized_name"],
        item["canonical_name"],

        item["category_name"],
        item["page_url"],

        item.get("inventory_image_url"),
        item.get("world_image_url"),
        item.get("color_image_url"),

        item.get("inventory_image_path"),
        item.get("world_image_path"),
        item.get("color_image_path"),

        item.get("internal_item_id"),
        item.get("internal_tile_id"),

        "strict_infobox",
        item["status"],
        item.get("problem"),
    ))

    connection.commit()


def download_available_images(all_items):
    for item in all_items:
        base = safe_filename(f"{item['local_id']}_{item['name']}")

        inventory_path, inventory_error = try_copy_image(
            item.get("inventory_image_url"),
            INVENTORY_IMAGES_DIR,
            base + "_inventory"
        )

        world_path, world_error = try_copy_image(
            item.get("world_image_url"),
            WORLD_IMAGES_DIR,
            base + "_world"
        )

        color_path, color_error = try_copy_image(
            item.get("color_image_url"),
            COLOR_IMAGES_DIR,
            base + "_color"
        )

        item["inventory_image_path"] = inventory_path
        item["world_image_path"] = world_path
        item["color_image_path"] = color_path

        if inventory_error and item.get("inventory_image_url"):
            item["problem"] = append_problem(item.get("problem"), "inventory_download_error")

        if world_error and item.get("world_image_url"):
            item["problem"] = append_problem(item.get("problem"), "world_download_error")

        if color_error and item.get("color_image_url"):
            item["problem"] = append_problem(item.get("problem"), "color_download_error")


def append_problem(existing, new_problem):
    if not existing:
        return new_problem

    if new_problem in existing:
        return existing

    return existing + "+" + new_problem


def get_missing_roles(item):
    missing = []

    if not item.get("inventory_image_url"):
        missing.append("inventory")

    if not item.get("world_image_url"):
        missing.append("world")

    # Color sera auto-rempli depuis world par manual_fix.
    return missing


def save_all_entries_csv(all_items):
    with ALL_ENTRIES_CSV_PATH.open("w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)

        writer.writerow([
            "local_id",
            "name",
            "canonical_name",
            "category_name",
            "status",
            "problem",
            "page_url",
            "inventory_image_url",
            "world_image_url",
            "color_image_url",
            "inventory_image_path",
            "world_image_path",
            "color_image_path",
            "internal_item_id",
            "internal_tile_id",
        ])

        for item in all_items:
            writer.writerow([
                item["local_id"],
                item["name"],
                item["canonical_name"],
                item["category_name"],
                item["status"],
                item.get("problem"),
                item["page_url"],
                item.get("inventory_image_url"),
                item.get("world_image_url"),
                item.get("color_image_url"),
                item.get("inventory_image_path"),
                item.get("world_image_path"),
                item.get("color_image_path"),
                item.get("internal_item_id"),
                item.get("internal_tile_id"),
            ])


def save_problem_entries_csv(problem_items):
    with PROBLEM_ENTRIES_CSV_PATH.open("w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)

        writer.writerow([
            "local_id",
            "name",
            "canonical_name",
            "category_name",
            "problem",
            "missing_roles",
            "page_url",
            "inventory_image_url",
            "world_image_url",
            "color_image_url",
            "internal_item_id",
            "internal_tile_id",
            "notes",
        ])

        for item in problem_items:
            writer.writerow([
                item["local_id"],
                item["name"],
                item["canonical_name"],
                item["category_name"],
                item.get("problem"),
                " | ".join(get_missing_roles(item)),
                item["page_url"],
                item.get("inventory_image_url"),
                item.get("world_image_url"),
                item.get("color_image_url"),
                item.get("internal_item_id"),
                item.get("internal_tile_id"),
                "",
            ])


def save_manual_fixes_csv(problem_items):
    """
    Fichier destiné au futur manual_fix.py.

    Tu rempliras seulement la colonne url.
    image_role :
    - inventory
    - world

    Si image_role = world, manual_fix.py remplira aussi color avec la même image.
    """
    with MANUAL_FIXES_CSV_PATH.open("w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)

        writer.writerow([
            "local_id",
            "name",
            "problem",
            "image_role",
            "url",
            "page_url",
            "notes",
        ])

        for item in problem_items:
            for role in get_missing_roles(item):
                writer.writerow([
                    item["local_id"],
                    item["name"],
                    item.get("problem"),
                    role,
                    "",
                    item["page_url"],
                    "",
                ])


def write_report(all_items, cache_hits, downloads, errors):
    ok_items = [item for item in all_items if item["status"] == "ok"]
    problem_items = [item for item in all_items if item["status"] == "problem"]

    lines = []

    lines.append("STRICT INFOBOX BUILD REPORT")
    lines.append("===========================")
    lines.append("")
    lines.append("RÈGLE")
    lines.append("-----")
    lines.append("Tous les blocs de l'Excel sont insérés dans la BDD.")
    lines.append("status=ok si inventory_image ET world_image sont trouvées dans l'infobox.")
    lines.append("status=problem si une info manque.")
    lines.append("Le local_id est attribué à tous les blocs, même incomplets.")
    lines.append("")
    lines.append("COUNTS")
    lines.append("------")
    lines.append(f"Total blocs en BDD : {len(all_items)}")
    lines.append(f"OK : {len(ok_items)}")
    lines.append(f"À compléter manuellement : {len(problem_items)}")
    lines.append(f"Erreurs Python/requête : {len(errors)}")
    lines.append("")
    lines.append("CACHE")
    lines.append("-----")
    lines.append(f"Pages HTML depuis cache : {cache_hits}")
    lines.append(f"Pages HTML téléchargées : {downloads}")
    lines.append("")
    lines.append("PROBLÈMES À CORRIGER")
    lines.append("--------------------")

    for item in problem_items:
        lines.append(
            f"[{item['local_id']}] {item['name']} | "
            f"missing={', '.join(get_missing_roles(item))} | "
            f"problem={item.get('problem')} | "
            f"page={item['page_url']}"
        )

    if errors:
        lines.append("")
        lines.append("ERREURS")
        lines.append("-------")

        for error in errors:
            lines.append(f"- {error}")

    lines.append("")
    lines.append("FICHIERS")
    lines.append("--------")
    lines.append(f"BDD : {DB_PATH}")
    lines.append(f"Excel lu : {EXCEL_READ_CSV_PATH}")
    lines.append(f"Toutes les entrées : {ALL_ENTRIES_CSV_PATH}")
    lines.append(f"Entrées problème : {PROBLEM_ENTRIES_CSV_PATH}")
    lines.append(f"Manual fixes : {MANUAL_FIXES_CSV_PATH}")
    lines.append(f"Rapport : {REPORT_PATH}")

    REPORT_PATH.write_text("\n".join(lines), encoding="utf-8")


def main():
    setup_dirs()

    print("Lecture Excel Blocks...")
    excel_entries = read_blocks_from_excel()
    write_excel_read_csv(excel_entries)

    print(f"Blocs lus : {len(excel_entries)}")
    print()

    all_items = []
    errors = []

    cache_hits = 0
    downloads = 0

    print("Lecture stricte des infobox...")
    print()

    for block in excel_entries:
        local_id = len(all_items)
        name = block["name"]

        print(f"[{local_id}] {name}")

        try:
            result = parse_page_strict(block)

            if result.get("from_cache"):
                cache_hits += 1
            else:
                downloads += 1

            status = "ok" if result.get("problem") is None else "problem"

            item = {
                "local_id": local_id,
                "name": name,
                "normalized_name": normalize_name(name),
                "canonical_name": block["canonical_name"],
                "category_name": block["category_name"],

                "page_url": result["page_url"],

                "inventory_image_url": result.get("inventory_image_url"),
                "world_image_url": result.get("world_image_url"),
                "color_image_url": result.get("color_image_url"),

                "inventory_image_path": None,
                "world_image_path": None,
                "color_image_path": None,

                "internal_item_id": result.get("internal_item_id"),
                "internal_tile_id": result.get("internal_tile_id"),

                "status": status,
                "problem": result.get("problem"),
                "infobox_image_count": result.get("infobox_image_count"),
            }

            all_items.append(item)

            if status == "ok":
                print("  OK : inventory + world trouvées")
            else:
                print(f"  PROBLEM : {item['problem']} | missing={', '.join(get_missing_roles(item))}")

        except Exception as error:
            page_url = page_url_from_name(name)

            item = {
                "local_id": local_id,
                "name": name,
                "normalized_name": normalize_name(name),
                "canonical_name": block["canonical_name"],
                "category_name": block["category_name"],

                "page_url": page_url,

                "inventory_image_url": None,
                "world_image_url": None,
                "color_image_url": None,

                "inventory_image_path": None,
                "world_image_path": None,
                "color_image_path": None,

                "internal_item_id": None,
                "internal_tile_id": None,

                "status": "problem",
                "problem": "exception",
                "infobox_image_count": None,
            }

            all_items.append(item)

            message = f"[{local_id}] {name} | {type(error).__name__} | {error}"
            errors.append(message)

            print(f"  ERROR : {error}")

    print()
    print("Téléchargement/copie des images disponibles...")
    download_available_images(all_items)

    print("Création BDD avec tous les blocs...")
    connection = create_database()

    for item in all_items:
        insert_object(connection, item)

    connection.close()

    problem_items = [item for item in all_items if item["status"] == "problem"]

    save_all_entries_csv(all_items)
    save_problem_entries_csv(problem_items)
    save_manual_fixes_csv(problem_items)

    write_report(
        all_items=all_items,
        cache_hits=cache_hits,
        downloads=downloads,
        errors=errors,
    )

    print()
    print("Terminé.")
    print(f"BDD : {DB_PATH}")
    print(f"Total blocs : {len(all_items)}")
    print(f"OK : {len([x for x in all_items if x['status'] == 'ok'])}")
    print(f"À compléter : {len(problem_items)}")
    print(f"Rapport : {REPORT_PATH}")
    print(f"Manual fixes : {MANUAL_FIXES_CSV_PATH}")


if __name__ == "__main__":
    main()